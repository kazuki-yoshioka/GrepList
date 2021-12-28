from types import CellType
from openpyxl.workbook import workbook
from Config import Config
import openpyxl
import os
from Utility import Utility
from OutputModel import OutputModel
import pandas as pd


class SearchExcelFile():
    """[summary] 対象のExcelファイルにおいて、検索文字が記載されているのかチェックします

    Returns:
        [type]: [description]
    """
    resultList = []
    config: Config = None
    searchColRange = 1000
    searchRowRange = 1000

    def __init__(self, _config: Config):
        self.resultList = []
        self.config = _config
        self.config.setCommentOutList()

        # 検索範囲をセット
        self.__setSearchRange()

    def __setSearchRange(self):
        """[summary] Excelのシートの検索範囲をセットする
        """
        try:
            self.searchColRange = int(
                self.config.config['SEARCH_SHEET_COL_RANGE'])
            self.searchRowRange = int(
                self.config.config['SEARCH_SHEET_ROW_RANGE'])
        except Exception as e:
            self.searchColRange = 1000
            self.searchRowRange = 1000

    def searchCharInFile(self, targetChar: str = '', path: str = ''):
        """[summary] 対象の文字列を検索する

        Args:
            targetChar (str, optional): [description]. Defaults to ''.
            path (str, optional): [description]. Defaults to ''.
        """
        ext = Utility.getExtention(path=path)
        if ext == '.xlsx':
            self.__searchByXlsx(targetChar=targetChar, path=path)

        elif ext == '.xls':
            self.__searchByXls(targetChar=targetChar, path=path)

    def __searchByXlsx(self, targetChar: str = '', path: str = ''):
        """[summary]対象のExcelファイル内で検索処理を実施
        対処の拡張子は「.xlsx」

        Args:
            targetChar (str, optional): [description]. Defaults to ''.
            検索対象文字
            path (str, optional): [description]. Defaults to ''.
            対象のExcelファイルパス
        """
        wb = openpyxl.load_workbook(path)
        sheetNameList = wb.sheetnames

        for sheetName in sheetNameList:
            # シート内で検索
            self.__serachInSheet(targetChar=targetChar,
                                 wb=wb, sheetName=sheetName, path=path)

    def __serachInSheet(self, targetChar: str = '', wb: workbook = None, sheetName: str = '', path=''):
        """[summary] シート内で対象の文字を検索

        Args:
            targetChar (str, optional): [description]. Defaults to ''.
            wb (workbook, optional): [description]. Defaults to None.
            sheetName (str, optional): [description]. Defaults to ''.
        """
        ws = wb[sheetName]

        for _row in range(1, 1000):
            for _col in range(1, 1000):
                _cell = ws.cell(row=_row, column=_col)

                # セルに対象の文字が含まれている場合
                if self.__includeTargetChar(targetChar=targetChar, value=_cell.value):
                    # リストに追加
                    self.resultList.append(self._SearchExcelFile__makeFileInfo(
                        targetChar=targetChar, _cell=_cell, path=path, sheetName=sheetName, _col=_col, _row=_row))

    def __includeTargetChar(self, targetChar='', value=None):
        """[summary] セルに対象の文字が入っているのかチェック

        Args:
            targetChar (str, optional): [description]. Defaults to ''.
            value (str, optional): [description]. Defaults to ''.

        Returns:
            [type]: [description]
        """
        if(value == None):
            return False

        if (targetChar in str(value)) is False:
            return False

        return True

    def __makeFileInfo(self, targetChar='', _cell=None, path: str = '', sheetName: str = '', _col: int = 0, _row: int = 0):
        """[summary] ファイル情報を作製

        Args:
            line (str, optional): [description]. Defaults to ''.
            path (str, optional): [description]. Defaults to ''.
            lineNo (int, optional): [description]. Defaults to 0.

        Returns:
            [type]: [description]
        """
        model = OutputModel()

        # grep対象文字
        model.targetChar = targetChar

        # ファイル名
        model.fileName = os.path.basename(path)

        # ファイルパス
        model.filePath = path

        # シート名
        model.sheet = sheetName

        # セル
        model.cell = self.__changeColNoForOutput(
            strColNo=str(_col)) + str(_row)

        # 行の文字列
        model.line = Utility.replaceNewLineCode(str(_cell.value))

        return model

    def __changeColNoForOutput(self, strColNo: str = ''):
        """[summary] 行NoをExcel行数の文字列に変更

        Args:
            strColNo (str, optional): [description]. Defaults to ''.

        Returns:
            [type]: [description]
        """
        if(strColNo == ''):
            return ''

        tmp = int(strColNo)
        rtnColStr = ''

        while True:
            q: int = tmp // 26

            # まだ割り算できる
            if q > 0:
                # 商を文字列に追加
                rtnColStr = rtnColStr + self.__convertColNoToCellno(q)
                tmp = q
            else:
                # 余りを文字列に追加
                mod = tmp % 26
                rtnColStr = rtnColStr + self.__convertColNoToCellno(mod)
                break

        return rtnColStr

    def __convertColNoToCellno(self, no: int = 0):
        """[summary] Excelの行数をセルの文字列に変更

        Args:
            no (int, optional): [description]. Defaults to 0.

        Returns:
            [type]: [description]　A ~ Z
        """
        if no == 0 or no > 26:
            return ''

        if no == 1:
            return 'A'
        elif no == 2:
            return 'B'
        elif no == 3:
            return 'C'
        elif no == 4:
            return 'D'
        elif no == 5:
            return 'E'
        elif no == 6:
            return 'F'
        elif no == 7:
            return 'G'
        elif no == 8:
            return 'H'
        elif no == 9:
            return 'I'
        elif no == 10:
            return 'J'
        elif no == 11:
            return 'K'
        elif no == 12:
            return 'L'
        elif no == 13:
            return 'M'
        elif no == 14:
            return 'N'
        elif no == 15:
            return 'O'
        elif no == 16:
            return 'P'
        elif no == 17:
            return 'Q'
        elif no == 18:
            return 'R'
        elif no == 19:
            return 'S'
        elif no == 20:
            return 'T'
        elif no == 21:
            return 'U'
        elif no == 22:
            return 'V'
        elif no == 23:
            return 'W'
        elif no == 24:
            return 'X'
        elif no == 25:
            return 'Y'
        elif no == 26:
            return 'Z'

        return ''

    def __searchByXls(self,  targetChar: str = '', path: str = ''):
        """[summary]対象のExcelファイルの拡張子が「xls」の場合
        拡張子を変更して「__searchByXlsx」（検索処理）を実行

        Args:
            targetChar (str, optional): [description]. Defaults to ''.
            検索対象文字
            path (str, optional): [description]. Defaults to ''.
            対象のExcelファイルパス
        """
        dir, ext = os.path.splitext(path)
        newParh = dir + '.xlsx'

        # 拡張子を変換
        df = pd.read_excel(path, header=None)
        df.to_excel(newParh, index=False, header=False)

        # 拡張子を変換したファイルで検索
        self.__searchByXlsx(targetChar=targetChar, path=newParh)

        # 変換したファイルを削除
        os.remove(newParh)
